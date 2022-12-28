# Python Imports
from openpyxl import load_workbook
from copy import deepcopy
# Own Imports
from utils import check_duplicated, check_previous_data, check_default_assignments, get_data_row, insert_data_row, print_log
from constants import DATA_ACTUAL, LOG, PROCESS, PATH, FILENAMES

# Loop trough /data/new/REAL.xlsx entries adding them to /data/old/REAL.xlsx into /data/final/REAL.xlsx file
def process_actual_data(): 
    # Add new lines from NEW_ACTUAL to OLD_ACTUAL
    old_actual_workbook = load_workbook(PATH.OLD.value + FILENAMES.ACTUAL.value)
    old_actual_sheet = old_actual_workbook.active
    new_actual_workbook = load_workbook(PATH.NEW.value + FILENAMES.ACTUAL.value)
    new_actual_sheet = new_actual_workbook.active
    # To check for missing tagetiks etc against old COMMITMENTS file
    old_commitments_wb = load_workbook(PATH.OLD.value + FILENAMES.COMMITMENT.value)
    old_commitments_sh = old_commitments_wb.active

    # Log
    print(f'\n{PROCESS.ACTUAL.value} Start processing {str(new_actual_sheet.max_row - 1)} rows...')

    # Loop trough rows in new actual file
    counter, counter_error, counter_warnings = 0, 0, 0
    max_rows_check = old_actual_sheet.max_row + 1
    for i in range(2, new_actual_sheet.max_row + 1):
        # get data from row
        data = get_data_row(PROCESS.ACTUAL, new_actual_sheet, i, deepcopy(DATA_ACTUAL))
        # check if co_number already exists in final_actual (only take into account original row number)
        if (check_duplicated(old_actual_sheet, max_rows_check, data["co_number"]["data"])):
            counter_error += 1
            print_log(PROCESS.ACTUAL, LOG.ERROR, data["po_number"]["data"], data["po_position"]["data"], data["coste"]["data"], 'Entry duplicated')
            continue
        # row relevante to insert
        counter += 1
        # check for previous information (first against old COMMITMENTS then ACTUAL)
        flag_updated, data = check_previous_data(PROCESS.COMMITMENT, old_commitments_sh, data)
        if (flag_updated):
            counter_warnings += 1
            log_entry = f'Entry updated from COMMITMENT ##' 
            for key in data:
                if (data[key]["updated"]):
                    log_entry += f'{key}#{data[key]["data"]}##'
            print_log(PROCESS.ACTUAL, LOG.WARNING, i, data["po_number"]["data"], data["po_position"]["data"], data["coste"]["data"], log_entry.strip())
        else:
            flag_updated, data = check_previous_data(PROCESS.ACTUAL, old_actual_sheet, data)
            if (flag_updated):
                counter_warnings += 1
                log_entry = f'Entry updated from ACTUAL ##' 
                for key in data:
                    if (data[key]["updated"]):
                        log_entry += f'{key}#{data[key]["data"]}##'
                print_log(PROCESS.ACTUAL, LOG.WARNING, i, data["po_number"]["data"], data["po_position"]["data"], data["coste"]["data"], log_entry.strip())
        # check tagetik stills blank
        if (data["tagetik"]["data"] == None):
            flag_updated, data = check_default_assignments(PROCESS.ACTUAL, data)
            if (flag_updated):
                counter_warnings += 1
                print_log(PROCESS.ACTUAL, LOG.WARNING, i, data["po_number"]["data"], data["po_position"]["data"], data["cost"]["data"], 'Tagetik default mapping update')
            else:
                counter_error += 1
                print_log(PROCESS.ACTUAL, LOG.ERROR, i, data["po_number"]["data"], data["po_position"]["data"], data["cost"]["data"], 'Tagetik not found')
        # insert data at the end of file
        insert_data_row(PROCESS.ACTUAL, old_actual_sheet, old_actual_sheet.max_row + 1, data)
    
    # New actual file saved in final path
    old_actual_workbook.save(PATH.FINAL.value + FILENAMES.ACTUAL.value)
    
    # Log
    print(f'{PROCESS.ACTUAL.value} Finished processing... {str(counter)} rows added (WARNINGS: {str(counter_warnings)} - ERRORS: {str(counter_error)})')
