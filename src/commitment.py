# Python Imports
from openpyxl import load_workbook
from copy import deepcopy
# Own Imports
from utils import check_previous_data, get_data_row, update_data_cell, print_log, insert_data_row
from constants import DATA_COMMITMENT, LOG, PROCESS, PATH, FILENAMES

# Loop trough /data/new/REAL.xlsx entries adding them to /data/old/REAL.xlsx into /data/final/REAL.xlsx file
def process_commitment_data(): 
    # New commitment file replaces old commitment file
    old_workbook = load_workbook(PATH.OLD.value + FILENAMES.COMMITMENT.value)
    old_sheet = old_workbook.active
    new_workbook = load_workbook(PATH.NEW.value + FILENAMES.COMMITMENT.value)
    new_sheet = new_workbook.active

    # Log
    print(f'\n{PROCESS.COMMITMENT.value} Start processing {str(new_sheet.max_row - 1)} rows...')
    
    # Loop trough rows in new commitment file
    counter, counter_error, counter_warnings = 0, 0, 0
    for i in range(2, new_sheet.max_row + 1):
        # get data from row
        data = get_data_row(PROCESS.COMMITMENT, new_sheet, i, deepcopy(DATA_COMMITMENT))
        # row relevante to insert
        counter += 1
        # check for previous information
        flag_updated, data = check_previous_data(PROCESS.COMMITMENT, old_sheet, data)
        if (flag_updated):
            counter_warnings += 1
            log_entry = f'Entry updated ##' 
            for key in data:
                if (data[key]["updated"]):    
                    update_data_cell(new_sheet, i, data[key]['column'], data[key]['data'])
                    log_entry += f'{key}#{data[key]["data"]}##'
            print_log(PROCESS.COMMITMENT, LOG.WARNING, i, data["po_number"]["data"], data["po_position"]["data"], data["comprometido"]["data"], log_entry.strip())
        # Update tipo entry
        tipo_entry = 'PED'
        if (data['po_number']['data'] == None or data['po_number']['data'] == ''):
            tipo_entry = 'FOR'       
        elif (str(data['po_number']['data']).startswith('21')):
            tipo_entry = 'SPD'
        update_data_cell(new_sheet, i, data['tipo']['column'], tipo_entry)
        # Update solicitante
        if (data['solicitante']['data'] != None):
            update_data_cell(new_sheet, i, data['solicitante']['column'], data['solicitante']['data'].replace('i:0#.w|acciona\\', ''))
        # Update proveedor
        if (data['proveedor']['data'] != None or data['proveedor']['data'] == ''):
            data['proveedor']['data'] = '-'
        # check tagetik stills blank and PO exists (error)
        if ((data["tagetik"]["data"] == None or data["tagetik"]["data"] == '') and data["ignorar"]["data"] != True):
            counter_error += 1
            print_log(PROCESS.COMMITMENT, LOG.ERROR, i, data["po_number"]["data"], data["po_position"]["data"], data["comprometido"]["data"], 'Tagetik not found')    
    
    # Add forecast entries from old files
    for i in range(2, old_sheet.max_row + 1):
        data = get_data_row(PROCESS.COMMITMENT, old_sheet, i, deepcopy(DATA_COMMITMENT))
        if (data['tipo']['data'] == 'FOR'):
            counter += 1
            insert_data_row(new_sheet, new_sheet.max_row + 1, data)
            
    # New actual file saved in final path
    new_sheet.title = 'COMPROMETIDO'
    new_workbook.save(PATH.FINAL.value + FILENAMES.COMMITMENT.value)

    # Log
    print(f'{PROCESS.COMMITMENT.value} Finished processing... {str(counter)} rows (vs {old_sheet.max_row + 1}) (WARNINGS: {str(counter_warnings)} - ERRORS: {str(counter_error)})')