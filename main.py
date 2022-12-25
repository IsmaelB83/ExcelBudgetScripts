# Imports
from openpyxl import load_workbook
import shutil

# CONSTANTS
OLD_PATH = "./data/old/"
NEW_PATH = "./data/new/"
FINAL_PATH = "./data/final/"
# FILE NAMES
COMMITMENT = "COMPROMETIDO.xlsx" 
ACTUAL = "REAL.xlsx"

# Checks if co_number already exists in final ACTUAL file (duplicated)
def check_duplicated (file_sheet, co_number):
    for i in range(2,  file_sheet.max_row + 1):
        po = new_actual_sh.cell(row=i, column=5).value
        position = new_actual_sh.cell(row=i, column=6).value
        amount = new_actual_sh.cell(row=i, column=15).value
        old_number = file_sheet.cell(row=i, column=22).value
        if (old_number == co_number):
            print("ERROR - Entry duplicated - " + str(po) + "-" + str(position) + " (amount " + str(amount) + ") with CO number " + str(co_number))
            return True
    return False   

# Checks if previous entry in ACTUAL file has different information for solicitante and tagetik. 
# In that case current extraction is udpated with old information, which is supossed to be validated in previous budget loads.
def check_previous_data (file_sheet, po_number, po_position, solicitante, tagetik):
    for i in range(2,  file_sheet.max_row + 1):
        old_number = file_sheet.cell(row=i, column=5).value
        old_position = file_sheet.cell(row=i, column=6).value
        old_solicitante = file_sheet.cell(row=i, column=20).value
        old_tagetik = file_sheet.cell(row=i, column=21).value
        if (old_number == po_number and old_position == po_position):
            if (old_solicitante != solicitante or old_tagetik != tagetik):
                print("WARNING - Tagetik Updated - " + str(po_number) + "-" + str(po_position) + " (" + tagetik + "/" + old_tagetik + ")")
                return old_solicitante, old_tagetik
    return solicitante, tagetik

# Add new lines from NEW_ACTUAL to OLD_ACTUAL
new_actual_wb = load_workbook(NEW_PATH + ACTUAL)
new_actual_sh = new_actual_wb.active
new_actual_sh_mr = new_actual_sh.max_row
new_actual_sh_mc = new_actual_sh.max_column

old_actual_wb = load_workbook(OLD_PATH + ACTUAL)
old_actual_sh = old_actual_wb.active
old_actual_sh_mr = old_actual_sh.max_row

# loop trough rows in new actual file
counter = 0
for i in range(2, new_actual_sh_mr + 1):
    # get relevant information from current row
    po_number = new_actual_sh.cell(row=i, column=5).value
    po_position = new_actual_sh.cell(row=i, column=6).value
    solicitante = new_actual_sh.cell(row=i, column=20).value
    tagetik = new_actual_sh.cell(row=i, column=21).value
    co_number = new_actual_sh.cell(row=i, column=22).value
    # check if co_number already exists in final_actual
    if (check_duplicated(old_actual_sh, co_number)):
        continue
    # row relevante to insert
    counter += 1
    # check for previous solicitante / tagetik
    solicitante, tagetik = check_previous_data(old_actual_sh, po_number, po_position, solicitante, tagetik)   
    # check tagetik stills blank and PO exists (error)
    if (tagetik == "" and po_number != ""):
        print("ERROR - Tagetik not found - " + str(po_number) + "-" + str(po_position))
    # loop trough columns of current row
    for j in range(1, new_actual_sh_mc + 1):
        # read cell value from Excel source file and write it in target file
        c = new_actual_sh.cell(row=i, column=j)        
        value = c.value
        if (c.column == 20):
            value = solicitante
        elif (c.column == 21):
            value = tagetik
        old_actual_sh.cell(row=old_actual_sh_mr + i - 1, column=j).value = value
 
# New actual file saved in final path
old_actual_wb.save(FINAL_PATH + ACTUAL)
print ('OK - ACTUAL FILE UPDATED - ' + str(counter) + " rows added (now " + str(old_actual_sh.max_row - 1)+ ")")
